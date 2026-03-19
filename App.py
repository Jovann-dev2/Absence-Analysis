import re
import sys
import streamlit as st
from pathlib import Path
import pandas as pd
import numpy as np
from io import BytesIO
from uuid import uuid4
import altair as alt

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import ColorScaleRule

st.set_page_config(page_title="Unavailability Insights", layout="wide")

st.title("🔎 Unavailability Insights")
st.write(
    "Upload an appropriate data file below. Unavailability insights will be generated, aggregated by a certain group. "
)

# ----------------------------
# Year selection (NEW)
# ----------------------------
DEFAULT_YEAR = 2026
year_str = st.text_input("📅 Year of interest", value=str(DEFAULT_YEAR), help="Enter a 4-digit year (e.g., 2026)")
try:
    selected_year = int(year_str)
    if selected_year < 1900 or selected_year > 2100:
        raise ValueError
except Exception:
    st.warning("Please enter a valid 4-digit year between 1900 and 2100. Using default year 2026.")
    selected_year = DEFAULT_YEAR

OUTPUT_CSV = f"{selected_year}_expiries_by_shaft.csv"

# If you know the exact file name, set it here (e.g., 'data.xlsx').
# Otherwise leave as None and the script will auto-detect if there is exactly one .xlsx in the folder.
xlsx_path = st.file_uploader("📤 Upload the appropriate XLSX data file", type=["xlsx"])

file_bytes: bytes | None = None
if xlsx_path is not None:
    # Read the uploaded file into bytes; this will be used as the cache key
    file_bytes = xlsx_path.getvalue()

# ======================================================
# 🆕 Critical skill designations — fill in your list here
# ======================================================
HARD_CODED_CRITICAL_SKILLS: list[str] = [
    # TODO: Replace/extend with your full official list:
    "Operator Rock Drill Single Handed UG",
    "Operator Winch UG",
    "Operator Loco",
    "Team Leader Production UG",
    "Shift Supervisor Production UG",
    "Miner Stoping",
    "Miner Development",
    "Miner General"
    # "Blaster UG",
    # ...
]

# ----------------------------
# Helpers
# ----------------------------
def normalize_text(s: str) -> str:
    """Lowercase and collapse non-alphanumeric to spaces to make matching robust."""
    return re.sub(r'[^a-z0-9]+', ' ', str(s).strip().lower()).strip()

def first_header_row(df: pd.DataFrame, min_non_empty: int = 6, max_scan: int = 50) -> int:
    """
    Find the first row that likely contains column names:
    - Has at least `min_non_empty` non-empty cells.
    - Scans up to `max_scan` rows.
    """
    scan_limit = min(len(df), max_scan)
    for i in range(scan_limit):
        row = df.iloc[i]
        non_empty = row.map(lambda x: pd.notna(x) and str(x).strip() != '').sum()
        if non_empty >= min_non_empty:
            return i
    return 0  # fallback

def read_sheet_with_header_detection(xlsx_obj, sheet_name: str) -> pd.DataFrame:
    """
    Read a sheet (xlsx_obj can be a Path, BytesIO, or pd.ExcelFile) without assuming header placement;
    detect header row and return the data.
    """
    tmp = pd.read_excel(xlsx_obj, sheet_name=sheet_name, header=None, engine='openpyxl')
    hdr_row = first_header_row(tmp)
    header = tmp.iloc[hdr_row].astype(str).tolist()
    df = tmp.iloc[hdr_row + 1:].copy()
    df.columns = header
    # Drop rows that are completely empty
    df = df.dropna(how='all')
    return df

def find_column(df: pd.DataFrame, targets: list[str]) -> str | None:
    """
    Find a column in df matching any of the normalized `targets`.
    Matching is exact on normalized text or substring match to be tolerant.
    """
    norm_targets = [normalize_text(t) for t in targets]
    norm_cols = {col: normalize_text(col) for col in df.columns}

    # Exact normalized match first
    for col, ncol in norm_cols.items():
        if ncol in norm_targets:
            return col

    # Then substring match
    for col, ncol in norm_cols.items():
        if any(t in ncol for t in norm_targets):
            return col

    return None

def parse_dates_series(series: pd.Series) -> pd.Series:
    if pd.api.types.is_datetime64_any_dtype(series):
        return series

    parsed = pd.to_datetime(series, errors='coerce')

    need_excel = parsed.isna()
    numeric = pd.to_numeric(series.where(need_excel), errors='coerce')
    mask_excel = need_excel & numeric.notna() & (numeric.between(20000, 60000))
    if mask_excel.any():
        excel_base = pd.Timestamp('1899-12-30')
        parsed.loc[mask_excel] = excel_base + pd.to_timedelta(numeric.loc[mask_excel], unit='D')

    return parsed

def classify_sheet_with_overrides(
    df: pd.DataFrame,
    sheet_name: str
) -> tuple[str | None, dict]:
    """
    FAST PATH:
      - If the sheet name exists in SHEET_COLUMN_MAP and required columns exist → use it.
    FALLBACK:
      - Use automatic detection (existing logic).
    """

    # ----------------------------
    # ✅ FAST PATH — hard coded
    # ----------------------------
    cfg = SHEET_COLUMN_MAP.get(sheet_name)

    if cfg:
        legal_type = cfg.get("legal_type")
        col_map: dict = {}

        missing = False
        for key, col_name in cfg.items():
            if key == "legal_type":
                continue
            if col_name not in df.columns:
                missing = True
                break
            col_map[key] = col_name

        if not missing:
            st.write(f"⚡ Using hard-coded layout for sheet '{sheet_name}'")
            return legal_type, col_map

        st.warning(
            f"⚠️ Sheet '{sheet_name}': hard-coded columns missing — falling back to auto detection."
        )

    # ----------------------------
    # 🐢 FALLBACK — automatic detection
    # ----------------------------
    group_candidates = ['group shaft name', 'shaft name', 'group name', 'shaft']
    cof_date_candidates = ['next examination date']
    wp_date_candidates = ['permit expiry date', 'permit expiry', 'expiry date']
    al_date_candidates = ['date of last leave', 'last leave date', 'last leave']

    group_col = find_column(df, group_candidates)
    cof_col = find_column(df, cof_date_candidates)
    wp_col  = find_column(df, wp_date_candidates)
    al_col  = find_column(df, al_date_candidates)

    if group_col and cof_col:
        return 'COF Expiry', {'group': group_col, 'date': cof_col}

    if group_col and al_col:
        return 'Annual Leave Expiry', {'group': group_col, 'last_leave': al_col}

    if group_col and wp_col and 'permit' in normalize_text(wp_col):
        return 'Work Permit Expiry', {'group': group_col, 'date': wp_col}

    return None, {}

def month_pivot_counts(df: pd.DataFrame, group_col: str, date_col: str, legal_type: str, year: int) -> pd.DataFrame:
    """Return counts per month (1–12) for the given legal_type and year, grouped by group_col."""
    dates = parse_dates_series(df[date_col])
    valid = df[dates.notna()].copy()
    valid['_exp_date'] = dates[dates.notna()]
    valid['_year'] = valid['_exp_date'].dt.year
    valid['_month'] = valid['_exp_date'].dt.month

    # Keep only the target year
    valid = valid[valid['_year'] == year]

    # Drop missing groups
    valid = valid[valid[group_col].notna() & (valid[group_col].astype(str).str.strip() != '')]

    if valid.empty:
        # Return empty table with correct structure
        idx = pd.Index([], name='Shaft')
        cols = list(range(1, 13))
        pivot = pd.DataFrame(0, index=idx, columns=cols, dtype=int)
    else:
        pivot = (
            valid.groupby([group_col, '_month'])
                 .size()
                 .unstack('_month')
                 .reindex(columns=range(1, 13), fill_value=0)
        )
        pivot.index.name = 'Shaft'

    # Add Legal Type level
    pivot['Legal Type'] = legal_type
    pivot = pivot.reset_index().set_index(['Shaft', 'Legal Type'])
    return pivot

def month_pivot_counts_annual_leave(df: pd.DataFrame, group_col: str, last_leave_col: str, legal_type: str, year: int) -> pd.DataFrame:
    """Compute expiry = last_leave + 18 months, then pivot counts by month for the given year."""
    last_leave = parse_dates_series(df[last_leave_col])
    valid = df[last_leave.notna()].copy()
    valid['_last_leave'] = last_leave[last_leave.notna()]
    # Add 18 months
    valid['_exp_date'] = valid['_last_leave'] + pd.DateOffset(months=17)
    valid['_year'] = valid['_exp_date'].dt.year
    valid['_month'] = valid['_exp_date'].dt.month

    # Keep only the target year
    valid = valid[valid['_year'] == year]

    # Drop missing groups
    valid = valid[valid[group_col].notna() & (valid[group_col].astype(str).str.strip() != '')]

    if valid.empty:
        idx = pd.Index([], name='Shaft')
        cols = list(range(1, 13))
        pivot = pd.DataFrame(0, index=idx, columns=cols, dtype=int)
    else:
        pivot = (
            valid.groupby([group_col, '_month'])
                 .size()
                 .unstack('_month')
                 .reindex(columns=range(1, 13), fill_value=0)
        )
        pivot.index.name = 'Shaft'

    pivot['Legal Type'] = legal_type
    pivot = pivot.reset_index().set_index(['Shaft', 'Legal Type'])
    return pivot

def discover_input_xlsx(explicit_path: str | None) -> Path:
    """Find the input xlsx file. If explicit_path is None, auto-detect if exactly one .xlsx exists."""
    if explicit_path:
        p = Path(explicit_path)
        if not p.exists():
            raise FileNotFoundError(f"File not found: {explicit_path}")
        return p

    xlsx_files = sorted(Path('.').glob('*.xlsx'))
    if len(xlsx_files) == 0:
        raise FileNotFoundError("No .xlsx file found in current directory.")
    if len(xlsx_files) > 1:
        files_list = '\n'.join([f" - {f.name}" for f in xlsx_files])
        raise FileNotFoundError(f"Multiple .xlsx files found. Please set INPUT_XLSX.\n{files_list}")
    return xlsx_files[0]

# ----------------------------
# Aggregation helper for custom shaft groupings
# ----------------------------
def aggregate_by_custom_groups(df: pd.DataFrame, mapping: dict[str, str], month_cols: list[str]) -> pd.DataFrame:
    """
    Aggregate rows by a custom mapping from Shaft -> GroupName.
    Unmapped shafts retain their original name. Aggregation preserves 'Legal Type'.
    """
    if df.empty:
        return df

    tmp = df.copy()
    tmp['Shaft'] = tmp['Shaft'].astype(str)
    # New grouped name if mapped, else keep original
    tmp['_ShaftGrouped'] = tmp['Shaft'].map(mapping).fillna(tmp['Shaft'])

    # Group and sum monthly numeric columns; recompute Total afterwards
    grouped = (
        tmp.drop(columns=['Shaft'])
           .rename(columns={'_ShaftGrouped': 'Shaft'})
           .groupby(['Shaft', 'Legal Type'], as_index=False)[month_cols].sum()
    )
    grouped['Total'] = grouped[month_cols].sum(axis=1)

    # Ensure column order and integer types
    grouped = grouped[['Shaft', 'Legal Type'] + month_cols + ['Total']]
    for c in month_cols + ['Total']:
        grouped[c] = pd.to_numeric(grouped[c], errors='coerce').fillna(0).astype(int)

    return grouped

# ======================================================
# 🆕 Helper to collect all Designation values across sheets
# ======================================================
@st.cache_data(show_spinner="🔎 Scanning designations…")
def get_all_designations(file_bytes: bytes) -> list[str]:
    xls = pd.ExcelFile(BytesIO(file_bytes), engine='openpyxl')
    seen: set[str] = set()
    for sheet in xls.sheet_names:
        try:
            raw = read_sheet_with_header_detection(xls, sheet)
            desig_col = find_column(raw, ['designation'])
            if desig_col:
                vals = (
                    raw[desig_col]
                    .dropna()
                    .astype(str)
                    .map(lambda s: s.strip())
                )
                seen.update([v for v in vals if v != ''])
        except Exception:
            continue
    out = sorted(seen)
    return out

# ======================================================
# ✅ HARD-CODED SHEET → COLUMN MAP (FAST PATH)
# Fill this in according to your workbook structure.
# ======================================================

SHEET_COLUMN_MAP: dict[str, dict] = {
    # Sheet name (exact match) : configuration
    
    "COF Register": {
        "legal_type": "COF Expiry",
        "group": "Group Shaft Name",
        "date": "Next Examination Date",
    },

    "Work Permits": {
        "legal_type": "Work Permit Expiry",
        "group": "Group Shaft Name",
        "date": "Permit Expiry Date",
    },

    "Annual Leave": {
        "legal_type": "Annual Leave Expiry",
        "group": "Group Shaft Name",
        "last_leave": "Date of Last Leave",
    },
}

# ======================================================
# 🆕 Build result with Designation filter
# ======================================================
@st.cache_data(show_spinner="🔄 Reading and aggregating workbook…")
def build_result(
    file_bytes: bytes,
    year: int,
    designation_filter_mode: str,
    critical_designations_selected: tuple[str, ...],  # tuple to make cache key stable
) -> tuple[pd.DataFrame, list[str]]:
    """
    Heavy step: open the workbook, optionally filter by Designation, classify sheets, build the monthly pivot, and return:
      - result: DataFrame ready for display/CSV
      - cols_out: list of month columns (names) + 'Total' for downstream grouping
    The cache key includes (file_bytes, year, designation_filter_mode, critical_designations_selected).
    """
    # Open the workbook once from bytes
    xls = pd.ExcelFile(BytesIO(file_bytes), engine='openpyxl')

    # Normalized critical set for filtering (only used if mode != Both)
    crit_norm_set = {normalize_text(x) for x in critical_designations_selected if str(x).strip() != ''}
    use_filter = designation_filter_mode in {"Critical only", "Non-critical only"}

    pivots: list[pd.DataFrame] = []

    # Process each sheet and classify
    for sheet in xls.sheet_names:
        try:
            raw = read_sheet_with_header_detection(xls, sheet)

            # --- 🆕 Apply Designation filtering if requested ---
            if use_filter:
                desig_col = find_column(raw, ['designation'])
                if desig_col:
                    desig_norm = raw[desig_col].astype(str).map(normalize_text)
                    if designation_filter_mode == "Critical only":
                        raw = raw[desig_norm.isin(crit_norm_set)]
                    elif designation_filter_mode == "Non-critical only":
                        raw = raw[~desig_norm.isin(crit_norm_set)]
                else:
                    # No Designation column in this sheet; skip filtering for it
                    pass

            # If a sheet becomes empty after filtering, skip it early
            if raw.empty:
                continue

            legal_type, cmap = classify_sheet_with_overrides(raw, sheet)
            if not legal_type:
                continue

            group_col = cmap.get('group')
            if legal_type in ('COF Expiry', 'Work Permit Expiry'):
                date_col = cmap.get('date')
                if date_col is None or group_col is None:
                    continue
                piv = month_pivot_counts(raw, group_col, date_col, legal_type, year)
                pivots.append(piv)

            elif legal_type == 'Annual Leave Expiry':
                last_leave_col = cmap.get('last_leave')
                if last_leave_col is None or group_col is None:
                    continue
                piv = month_pivot_counts_annual_leave(raw, group_col, last_leave_col, legal_type, year)
                pivots.append(piv)
        except Exception as e:
            print(f"[WARN] Skipping sheet '{sheet}': {e}", file=sys.stderr)
            continue

    if not pivots:
        month_cols = list(range(1, 13))
        result = pd.DataFrame(columns=['Shaft', 'Legal Type'] + month_cols + ['Total'])
        # Keep a consistent month name list for downstream grouping
        cols_out = ['January','February','March','April','May','June',
                    'July','August','September','October','November','December','Total']
        return result, cols_out

    # Combine all pivots
    combined = pd.concat(pivots, axis=0)
    combined = combined.groupby(level=[0, 1]).sum()

    for m in range(1, 13):
        if m not in combined.columns:
            combined[m] = 0

    combined = combined.reindex(columns=list(range(1, 13)), fill_value=0)
    combined['Total'] = combined.sum(axis=1)

    month_map = {
        1:'January', 2:'February', 3:'March', 4:'April', 5:'May', 6:'June',
        7:'July', 8:'August', 9:'September', 10:'October', 11:'November', 12:'December'
    }
    combined = combined.rename(columns=month_map)

    cols_out = ['January','February','March','April','May','June',
                'July','August','September','October','November','December','Total']

    result = combined.reset_index()
    result = result[['Shaft', 'Legal Type'] + cols_out]

    for c in cols_out:
        result[c] = pd.to_numeric(result[c], errors='coerce').fillna(0).astype(int)

    return result, cols_out

def build_legals_xlsx_bytes(result_like_df: pd.DataFrame, cols_out: list[str], year: int) -> bytes:
    """
    Create an Excel file with structure like the provided screenshot:
      - Title "Legals {year}"
      - Columns: Shafts | Legal type | January..December | Total
      - For each Shaft: rows for COF / Work permit / Annual leave (Training excluded)
      - A "Total planned Expiries" row after each shaft block
      - Conditional formatting (reds) over monthly values
    Returns workbook bytes.
    """
    # Defensive copy and typing
    df = result_like_df.copy()
    if df.empty:
        # Build an empty template anyway
        df = pd.DataFrame(columns=['Shaft', 'Legal Type'] + cols_out)

    # Normalize and *exclude training* no matter how it appears
    def norm(s): 
        return str(s).strip().lower()

    df['__lt_norm'] = df['Legal Type'].apply(norm)
    df = df[~df['__lt_norm'].str.contains('train')]  # ignore anything that looks like Training

    # Keep only columns we need in the right order
    months = cols_out[:-1]  # months only (no 'Total')
    total_name = cols_out[-1]
    use_cols = ['Shaft', 'Legal Type'] + months + [total_name]
    df = df.reindex(columns=use_cols)

    # Make sure numeric
    for c in months + [total_name]:
        df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0).astype(int)

    # Map your data's "Legal Type" values to the display required in the sheet
    # Your code uses: 'COF Expiry', 'Work Permit Expiry', 'Annual Leave Expiry'
    display_map = {
        'cof expiry': 'COF',
        'work permit expiry': 'Work permit',
        'annual leave expiry': 'Annual leave',
    }
    df['__lt_display'] = df['Legal Type'].apply(lambda s: display_map.get(norm(s), s))

    # We will only include COF / Work permit / Annual leave in this order
    order_norm = ['cof expiry', 'work permit expiry', 'annual leave expiry']

    # ------------------- Build workbook -------------------
    wb = Workbook()
    ws = wb.active
    ws.title = f"Legals {year}"

    # Title (merge across B to last column)
    header_labels = ['Shafts', 'Legal type'] + months + [total_name]
    last_col_idx = 1 + len(header_labels)  # 1-based
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=last_col_idx)
    ws.cell(row=1, column=2).value = f"Legals {year}"
    ws.cell(row=1, column=2).font = Font(size=14, bold=True)
    ws.cell(row=1, column=2).alignment = Alignment(horizontal="center")

    # Header row
    header_row = 2
    for j, label in enumerate(header_labels, start=1):
        cell = ws.cell(row=header_row, column=j, value=label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    ws.column_dimensions['A'].width = 26  # Shafts
    ws.column_dimensions['B'].width = 16  # Legal type
    for ci in range(3, last_col_idx + 1):
        # best-effort; wide enough by default
        pass

    # Freeze panes below header and to the right of "Legal type"
    ws.freeze_panes = "C3"

    # Borders
    thin = Side(border_style="thin", color="DDDDDD")

    # Fill rows
    row_cursor = header_row + 1

    # Iterate shafts in their current order of appearance
    for shaft in df['Shaft'].astype(str).fillna('').unique():
        df_shaft = df[df['Shaft'].astype(str) == shaft]

        # Write 3 rows in the order (COF, Work permit, Annual leave) — Training is NOT included
        first_row_for_shaft = True
        for lt_norm in order_norm:
            sub = df_shaft[df_shaft['Legal Type'].str.lower() == lt_norm]

            # If a type is missing for a shaft, use zeros
            if sub.empty:
                values = [0 for _ in months] + [0]
                lt_display = display_map.get(lt_norm, lt_norm.title())
            else:
                # If multiple rows somehow exist (after grouping etc.), sum them
                summed = sub[months + [total_name]].sum(axis=0)
                values = [int(summed[m]) for m in months] + [int(summed[total_name])]
                lt_display = sub['__lt_display'].iloc[0]

            # Shaft name only on the first row of the block
            ws.cell(row=row_cursor, column=1, value=shaft if first_row_for_shaft else "")
            ws.cell(row=row_cursor, column=2, value=lt_display)

            # Write months + total
            for j, m in enumerate(months + [total_name], start=3):
                ws.cell(row=row_cursor, column=j, value=values[j - 3])
                ws.cell(row=row_cursor, column=j).number_format = '0'

            # Style row borders lightly
            for j in range(1, last_col_idx + 1):
                ws.cell(row=row_cursor, column=j).border = Border(top=thin, bottom=thin, left=thin, right=thin)

            first_row_for_shaft = False
            row_cursor += 1

        # Totals row for this shaft (sum across the three rows we just wrote)
        # Compute directly from df_shaft (safer if some types were absent)
        tot_series = (
            df_shaft[df_shaft['Legal Type'].str.lower().isin(order_norm)][months + [total_name]]
            .sum(axis=0)
        )
        tot_vals = [int(tot_series[m]) for m in months] + [int(tot_series[total_name])]

        ws.cell(row=row_cursor, column=1, value="Total planned Expiries")
        ws.cell(row=row_cursor, column=1).font = Font(bold=True)
        ws.cell(row=row_cursor, column=2, value="")  # empty "Legal type" cell

        for j, m in enumerate(months + [total_name], start=3):
            ws.cell(row=row_cursor, column=j, value=tot_vals[j - 3])
            ws.cell(row=row_cursor, column=j).font = Font(bold=True)
            ws.cell(row=row_cursor, column=j).number_format = '0'

        # Light fill for the totals row (optional)
        for j in range(1, last_col_idx + 1):
            ws.cell(row=row_cursor, column=j).fill = PatternFill("solid", fgColor="F4F4F5")
            ws.cell(row=row_cursor, column=j).border = Border(top=thin, bottom=thin, left=thin, right=thin)

        row_cursor += 1  # next block starts after this row

    last_data_row = row_cursor - 1

    # Conditional formatting (reds) across month cells (C3 : last_col x last_data_row)
    if last_data_row >= 3:
        cf_range = f"C3:{ws.cell(row=last_data_row, column=last_col_idx).coordinate}"
        ws.conditional_formatting.add(
            cf_range,
            ColorScaleRule(
                start_type='min', start_color='FFFAFA',   # very light
                mid_type='percentile', mid_value=50, mid_color='FCA5A5',  # light red
                end_type='max', end_color='DC2626'  # deep red
            )
        )

    # Pack to bytes
    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

# ----------------------------
# Main
# ----------------------------
def main(file_bytes: bytes | None):
    if not file_bytes:
        st.info("Please upload an XLSX file to begin.")
        return

    # ======================================================
    # 🆕 UI: Designation (Critical Skills) filter
    # ======================================================
    st.subheader("🎯 Designation Filter (Critical Skills)")
    mode = st.selectbox(
        "Which designations should be included?",
        options=["Both (no filter)", "Critical only", "Non-critical only"],
        index=0,
        help="Apply a filter based on the 'Designation' column if present in your sheets."
    )

    # Collect all dataset designations to drive UI/defaults
    dataset_designations = get_all_designations(file_bytes)

    # Defaults: any hard-coded critical designations that are present in the dataset
    defaults_from_hardcoded = sorted(
        set(dataset_designations) & set(HARD_CODED_CRITICAL_SKILLS)
    )

    selected_critical: list[str] = []
    if mode != "Both (no filter)":
        st.caption("Select the **critical designations** below. Defaults include your hard-coded items that are present in the data.")
        selected_critical = st.multiselect(
            "Critical designations",
            options=dataset_designations,
            default=defaults_from_hardcoded,
            help="These will be treated as 'critical'."
        )
        extra = st.text_input(
            "Add additional critical designations (comma/semicolon/newline separated, optional)",
            value="",
            placeholder="e.g., Specialist XYZ; Senior Blaster UG"
        )
        if extra.strip():
            extras = [s.strip() for s in re.split(r'[,\n;]+', extra) if s.strip()]
            selected_critical = sorted(set(selected_critical) | set(extras))

    # Build a cache-stable tuple for the selected critical designations
    selected_critical_tuple = tuple(sorted(set(selected_critical), key=lambda s: s.lower()))

    # Map display text to internal mode values
    mode_internal = {
        "Both (no filter)": "Both",
        "Critical only": "Critical only",
        "Non-critical only": "Non-critical only",
    }[mode]

    # Use the cached heavy computation (now includes designation filter parameters)
    result, cols_out = build_result(
        file_bytes=file_bytes,
        year=selected_year,
        designation_filter_mode=mode_internal,
        critical_designations_selected=selected_critical_tuple
    )

    # Save and show the base (un-grouped) result
    result.to_csv(OUTPUT_CSV, index=False)
    st.dataframe(result, width='stretch')

    # ----------------------------
    # Optional custom groupings UI (progressive add/remove)
    # ----------------------------
    st.subheader("🧩 Optional: Group 'Shaft' names into custom groups")
    grouping_choice = st.radio(
        "Do you want to form groupings of shaft names (for comprehensiveness)?",
        options=["No", "Yes"], horizontal=True
    )

    result_to_show = result.copy()
    output_name = OUTPUT_CSV

    if grouping_choice == "Yes" and not result.empty:
        shafts = sorted(result['Shaft'].astype(str).dropna().unique().tolist())

        # Reset group definitions when the uploaded data (shafts list) changes
        if "shafts_signature" not in st.session_state or st.session_state.shafts_signature != tuple(shafts):
            st.session_state.shafts_signature = tuple(shafts)
            st.session_state.group_defs = []  # clear any stale group definitions

        # Initialize the container for group definitions in session state
        if "group_defs" not in st.session_state:
            st.session_state.group_defs = []

        def add_group():
            """Append a new empty group with a unique ID."""
            st.session_state.group_defs.append({
                "id": str(uuid4()),
                "name": f"Group {len(st.session_state.group_defs) + 1}",
                "members": []
            })

        st.caption("Create a group, then add as many more as you need. You can also remove groups.")

        exclude_ungrouped = st.checkbox(
            "Exclude ungrouped shafts",
            value=False,
            help="When enabled, only shafts that are assigned to a group will be included in the aggregated results."
        )

        # If there are no groups yet, offer to create the first one
        if not st.session_state.group_defs:
            st.info("No groups yet. Click below to create your first group.")
            st.button("➕ Add first group", on_click=add_group)

        # Render all groups (if any)
        to_remove_idx = None
        for idx, g in enumerate(st.session_state.group_defs):
            with st.expander(f"Group {idx + 1} configuration", expanded=(idx == 0)):
                # Editable group name
                name = st.text_input(
                    f"Name for group {idx + 1}",
                    value=g["name"],
                    key=f"group_name_{g['id']}"
                )

                # Members selection
                label_for_members = f"Select shaft names for '{name or g['name']}'"
                members = st.multiselect(
                    label_for_members,
                    options=shafts,
                    default=g["members"],
                    key=f"group_members_{g['id']}"
                )

                # Persist current values back to session state
                g["name"] = name
                g["members"] = members

                # Row of small action buttons for this group
                c1, c2, _ = st.columns([1, 1, 6])
                with c1:
                    if st.button("🗑️ Remove group", key=f"remove_group_{g['id']}"):
                        to_remove_idx = idx
                with c2:
                    # Quick-add a new group right after this one
                    if st.button("➕ Add another group", key=f"add_group_after_{g['id']}"):
                        add_group()

        # Apply removal (if requested) and rerun to keep keys stable
        if to_remove_idx is not None:
            st.session_state.group_defs.pop(to_remove_idx)
            st.rerun()

        # Global add button at the bottom (available once at least one group exists)
        if st.session_state.group_defs:
            st.button("➕ Add another group", on_click=add_group, key="add_group_bottom")

        # --- Build mapping & detect conflicts across all groups ---
        conflict_tracker: dict[str, str] = {}
        for g in st.session_state.group_defs:
            group_name = g["name"] or "Unnamed Group"
            for s in g["members"]:
                if s in conflict_tracker and conflict_tracker[s] != group_name:
                    conflict_tracker[s] = "__CONFLICT__"
                else:
                    conflict_tracker[s] = group_name

        conflicts = [s for s, v in conflict_tracker.items() if v == "__CONFLICT__"]
        if conflicts:
            st.error(
                "The following shafts are assigned to multiple groups. "
                "Please resolve the overlaps:\n\n- " + "\n- ".join(sorted(conflicts))
            )
        else:
            # mapping = shaft -> group_name
            mapping = {s: grp for s, grp in conflict_tracker.items()}

            if mapping:
                # If exclude_ungrouped is enabled, keep only rows whose Shaft is mapped
                if exclude_ungrouped:
                    filtered = result[result["Shaft"].astype(str).isin(mapping.keys())].copy()
                    # cols_out includes 'Total' at the end; monthly columns are cols_out[:-1]
                    result_to_show = aggregate_by_custom_groups(filtered, mapping, month_cols=cols_out[:-1])
                    output_name = OUTPUT_CSV.replace(".csv", "_grouped_exclusive.csv")
                    st.success("✅ Custom grouping applied (ungrouped shafts excluded). The table below shows the aggregated results.")
                else:
                    result_to_show = aggregate_by_custom_groups(result, mapping, month_cols=cols_out[:-1])
                    output_name = OUTPUT_CSV.replace(".csv", "_grouped.csv")
                    st.success("✅ Custom grouping applied. The table below shows the aggregated results.")
            else:
                if exclude_ungrouped:
                    st.info("No shafts were assigned to any group. With exclusion enabled, nothing to show. Disable exclusion or add groups.")
                    # Show an empty frame with the same columns to avoid errors
                    result_to_show = result.head(0).copy()
                else:
                    st.info("No shafts were assigned to any group. Showing the original results.")

    # ----------------------------
    # Save & display
    # ----------------------------
    st.dataframe(result_to_show, width='stretch')

    # ----------------------------
    # 🔥 Heat map of the final table (NEW)
    # ----------------------------
    st.subheader("🔥 Monthly Heat Map per Shaft Group")
    if result_to_show.empty:
        st.info("No data available to display the heat map.")
    else:
        month_names = cols_out[:-1]  # exclude 'Total'
        # Tidy format
        tidy = result_to_show.melt(
            id_vars=['Shaft', 'Legal Type', 'Total'],
            value_vars=month_names,
            var_name='Month',
            value_name='Count'
        )
        tidy['Month'] = pd.Categorical(tidy['Month'], categories=month_names, ordered=True)
        tidy['Row Label'] = tidy['Shaft'].astype(str) + " – " + tidy['Legal Type'].astype(str)

        heat = (
            alt.Chart(tidy)
            .mark_rect()
            .encode(
                x=alt.X('Month:O', sort=month_names, title="Month"),
                y=alt.Y('Row Label:N', title="Shaft – Legal Type"),
                color=alt.Color('Count:Q', title="Count", scale=alt.Scale(scheme='reds')),
                tooltip=[
                    alt.Tooltip('Shaft:N'),
                    alt.Tooltip('Legal Type:N'),
                    alt.Tooltip('Month:N'),
                    alt.Tooltip('Count:Q')
                ]
            )
            .properties(height=min(28 * max(1, tidy['Row Label'].nunique()), 800), width='container')
        )
        st.altair_chart(heat, width='stretch')

    # ----------------------------
    # 📈 Time series plot for chosen year (NEW)
    # ----------------------------
    st.subheader(f"📈 Time Series – {selected_year}")
    if result_to_show.empty:
        st.info("No data available to display the time series.")
    else:
        month_names = cols_out[:-1]
        tidy = result_to_show.melt(
            id_vars=['Shaft', 'Legal Type', 'Total'],
            value_vars=month_names,
            var_name='Month',
            value_name='Count'
        )
        tidy['Month'] = pd.Categorical(tidy['Month'], categories=month_names, ordered=True)
        tidy['Series'] = tidy['Shaft'].astype(str) + " – " + tidy['Legal Type'].astype(str)

        shafts_all = sorted(result_to_show['Shaft'].astype(str).unique().tolist())
        # Default to top 5 shafts by total for readability
        top_by_total = (
            result_to_show[['Shaft', 'Total']]
            .groupby('Shaft', as_index=False).sum()
            .sort_values('Total', ascending=False)['Shaft']
            .astype(str)
            .tolist()
        )
        default_shafts = top_by_total[:5] if top_by_total else []

        legal_all = sorted(result_to_show['Legal Type'].astype(str).unique().tolist())

        c_opts, c_chart = st.columns([1, 2], gap="medium")
        with c_opts:
            selected_shafts = st.multiselect(
                "Groups/Shafts to plot",
                options=shafts_all,
                default=default_shafts,
                help="Select which groups/shafts to show on the chart."
            )
            selected_legal = st.multiselect(
                "Expiry type(s)",
                options=legal_all,
                default=legal_all,
                help="Filter by type of expiry (e.g., COF, Work Permit, Annual Leave)."
            )

        # Filter and build series
        filtered = tidy[
            tidy['Shaft'].astype(str).isin(selected_shafts) &
            tidy['Legal Type'].astype(str).isin(selected_legal)
        ].copy()

        if filtered.empty:
            c_chart.info("Adjust your filters—there’s no data for the current selection.")
        else:
            line = (
                alt.Chart(filtered)
                .mark_line(point=True)
                .encode(
                    x=alt.X('Month:O', sort=month_names, title="Month"),
                    y=alt.Y('Count:Q', title="Monthly Count"),
                    color=alt.Color('Series:N', title="Series (Shaft – Legal Type)"),
                    tooltip=[
                        alt.Tooltip('Shaft:N'),
                        alt.Tooltip('Legal Type:N'),
                        alt.Tooltip('Month:N'),
                        alt.Tooltip('Count:Q')
                    ]
                )
                .properties(width='container', height=420)
            )
            c_chart.altair_chart(line, width='stretch')

    # ======================================================
    # 🟥 NEW SECTION — TOTAL EXPIRIES (ALL LEGAL TYPES)
    # ======================================================

    st.header("📊 Total Expiries per Month (Summed Across All Legal Types)")

    if result_to_show.empty:
        st.info("No data available to calculate total expiries.")
    else:
        # ----------------------------
        # Aggregate totals per Shaft/Group
        # ----------------------------
        month_names = cols_out[:-1]  # Months only, not "Total"
        totals_only = (
            result_to_show
            .groupby("Shaft")[month_names]
            .sum()
            .reset_index()
        )

        st.subheader("📋 Totals Table")
        st.dataframe(totals_only, width='stretch')

        # ----------------------------
        # 🔥 Heat Map (Totals Only)
        # ----------------------------
        st.subheader("🔥 Risk Heat Map (Totals Only)")

        tidy_totals = totals_only.melt(
            id_vars="Shaft",
            value_vars=month_names,
            var_name="Month",
            value_name="Count"
        )
        tidy_totals["Month"] = pd.Categorical(tidy_totals["Month"], categories=month_names, ordered=True)

        heat_tot = (
            alt.Chart(tidy_totals)
            .mark_rect()
            .encode(
                x=alt.X("Month:O", sort=month_names, title="Month"),
                y=alt.Y("Shaft:N", title="Shaft/Group"),
                color=alt.Color("Count:Q", scale=alt.Scale(scheme="reds"), title="Total Expiries"),
                tooltip=["Shaft", "Month", "Count"]
            )
            .properties(
                width="container",
                height=min(30 * len(totals_only), 800)
            )
        )

        st.altair_chart(heat_tot, width='stretch')

        # ----------------------------
        # 📈 Time Series (Totals Only)
        # ----------------------------
        st.subheader("📈 Time Series of Total Expiries")

        shafts_available = sorted(totals_only["Shaft"].unique().tolist())
        default_shaft_selection = shafts_available[:5]

        left_filt, right_plot = st.columns([1, 2])

        with left_filt:
            selected_shaft_totals = st.multiselect(
                "Select Shaft(s)/Group(s)",
                options=shafts_available,
                default=default_shaft_selection,
            )

        if not selected_shaft_totals:
            right_plot.info("Please select at least one shaft.")
        else:
            filtered = tidy_totals[tidy_totals["Shaft"].isin(selected_shaft_totals)]

            line_tot = (
                alt.Chart(filtered)
                .mark_line(point=True)
                .encode(
                    x=alt.X("Month:O", sort=month_names),
                    y=alt.Y("Count:Q", title="Total Expiries"),
                    color=alt.Color("Shaft:N", title="Shaft/Group"),
                    tooltip=["Shaft", "Month", "Count"]
                )
                .properties(width="container", height=420)
            )

            right_plot.altair_chart(line_tot, width='stretch')

    # =============================
    # ⬇️ Download: Excel like the screenshot (ignore Training)
    # =============================
    st.subheader("⬇️ Download Excel — Legals layout")
    if result_to_show.empty:
        st.info("No data available to export.")
    else:
        try:
            xlsx_bytes = build_legals_xlsx_bytes(result_to_show, cols_out, selected_year)
            st.download_button(
                label=f"Download Legals {selected_year} (XLSX)",
                data=xlsx_bytes,
                file_name=f"Legals_{selected_year}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch'
            )
        except Exception as ex:
            st.error(f"Could not build the Excel file: {ex}")

if __name__ == "__main__":
    main(file_bytes)
